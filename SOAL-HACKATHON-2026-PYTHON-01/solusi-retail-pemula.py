# ==========================================================================
# SOLUSI HACKATHON RETAIL - VERSI PEMULA (SKOR 100)
# ==========================================================================
#
# File ini adalah versi yang mudah dipahami untuk pemula Python.
# Logika dan hasilnya SAMA PERSIS dengan versi advanced (Skor 100).
#
# Setiap langkah diberi komentar penjelasan agar mudah dipelajari.
# ==========================================================================

# --------------------------------------------------------------------------
# LANGKAH 0: IMPORT LIBRARY
# --------------------------------------------------------------------------
# matplotlib.use('Agg') = pakai mode tanpa tampilan grafik (untuk server Linux)
# warnings = agar pesan peringatan tidak mengganggu output kita
# pandas = library utama untuk olah data tabel
# numpy = library untuk perhitungan matematika
# matplotlib = library untuk membuat grafik
# mlxtend = library untuk algoritma Apriori (analisis keranjang belanja)
# openpyxl = library untuk mengedit file Excel hasil ekspor

import matplotlib
matplotlib.use('Agg')  # Wajib di awal sebelum import plt

import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from mlxtend.frequent_patterns import apriori, association_rules
from mlxtend.preprocessing import TransactionEncoder
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ==========================================================================
# BAGIAN 1: MUAT DATA
# ==========================================================================

print("Memuat data...")

# Baca file Excel, ambil sheet bernama 'Transaksi'
df = pd.read_excel('data_penjualan.xlsx', sheet_name='Transaksi')

# Ubah kolom tanggal jadi tipe datetime agar bisa diurutkan dengan benar
df['tgl_transaksi'] = pd.to_datetime(df['tgl_transaksi'])

print(f"Data berhasil dimuat: {len(df)} baris transaksi")


# ==========================================================================
# BAGIAN 2: ANALISIS RISING STAR
# ==========================================================================
# Rising Star = produk yang Moving Average (MA) penjualannya naik terus.
#
# Langkah-langkahnya:
#   1. Hitung total penjualan HARIAN per produk (kelompokkan by kode & nama)
#   2. Hitung Moving Average 3 hari (tanpa min_periods agar 2 hari awal NaN)
#   3. Cari semua sesi tren kenaikan MA berturut-turut >= 12 hari
#   4. Hitung growth di dalam sesi tren tersebut, pilih yang tertinggi per produk
# ==========================================================================

print("Menganalisis Rising Star...")

# --------------------------------------------------------------------------
# LANGKAH 2.1: Hitung total penjualan HARIAN per produk
# --------------------------------------------------------------------------
# Kita kelompokkan berdasarkan kode_produk, nama_produk, dan tgl_transaksi.
# Pengelompokan dengan kode_produk lebih aman jika ada nama produk sama.

daily_sales = (
    df.groupby(['kode_produk', 'nama_produk', 'tgl_transaksi'])['total_nilai']
    .sum()
    .reset_index()
)
daily_sales = daily_sales.sort_values(['kode_produk', 'tgl_transaksi']).reset_index(drop=True)

# --------------------------------------------------------------------------
# LANGKAH 2.2: Hitung Moving Average (MA) 3 hari
# --------------------------------------------------------------------------
# Sesuai logika juri, inisialisasi default MA 3 hari (window=3, min_periods=3)
# digunakan, sehingga 2 hari pertama untuk setiap produk akan bernilai NaN.

daftar_produk_codes = daily_sales['kode_produk'].unique()
daily_sales['MA'] = 0.0

for kode in daftar_produk_codes:
    mask = daily_sales['kode_produk'] == kode
    nilai_harian = daily_sales.loc[mask, 'total_nilai']
    
    # Rata-rata bergerak 3 hari (default min_periods=3)
    ma_hasil = nilai_harian.rolling(window=3).mean()
    daily_sales.loc[mask, 'MA'] = ma_hasil

# --------------------------------------------------------------------------
# LANGKAH 2.3 & 2.4: Cari semua sesi tren naik >= 12 hari & hitung growth
# --------------------------------------------------------------------------
# Logika Juri: Kita menguji SEMUA sesi kenaikan beruntun >= 12 hari pada suatu produk,
# lalu mengambil sesi yang menghasilkan persentase Growth % tertinggi.

def find_best_rising_session(group):
    group = group.sort_values('tgl_transaksi').reset_index(drop=True)
    ma = group['MA'].values

    streaks = []
    current_streak = 0

    # Deteksi seluruh sesi kenaikan MA berturut-turut
    for i in range(1, len(ma)):
        if pd.notna(ma[i]) and pd.notna(ma[i-1]) and ma[i] > ma[i-1]:
            current_streak += 1
        else:
            # Jika tren putus dan sebelumnya naik >= 12 hari, catat sesinya
            if current_streak >= 12:
                streaks.append({
                    'streak': current_streak,
                    'start_idx': i - current_streak,
                    'end_idx': i - 1
                })
            current_streak = 0

    # Cek apakah sesi terakhir berlanjut hingga akhir data
    if current_streak >= 12:
        streaks.append({
            'streak': current_streak,
            'start_idx': len(ma) - current_streak,
            'end_idx': len(ma) - 1
        })

    if not streaks:
        return None

    # Cari sesi dengan Growth % terbesar
    best_session = None
    max_growth_pct = -float('inf')

    for s in streaks:
        ma_start = ma[s['start_idx']]
        ma_end = ma[s['end_idx']]
        
        # Hitung pertumbuhan persentase
        growth_pct = ((ma_end - ma_start) / ma_start) * 100
        
        if growth_pct > max_growth_pct:
            max_growth_pct = growth_pct
            best_session = {
                'kode_produk': group['kode_produk'].iloc[0],
                'max_consecutive_days': s['streak'],
                'growth_pct': round(growth_pct, 2)
            }

    return best_session

rising_results = []
for kode, group in daily_sales.groupby('kode_produk'):
    result = find_best_rising_session(group)
    if result is not None:
        rising_results.append(result)

rising_stars_df = pd.DataFrame(rising_results)

# --------------------------------------------------------------------------
# LANGKAH 2.5: Gabungkan dengan nama produk & total penjualan
# --------------------------------------------------------------------------
product_totals = (
    df.groupby(['kode_produk', 'nama_produk'])['total_nilai']
    .sum()
    .reset_index()
    .rename(columns={'total_nilai': 'total_penjualan'})
)

final_report = rising_stars_df.merge(product_totals, on='kode_produk', how='left')
final_report = final_report.sort_values('growth_pct', ascending=False).reset_index(drop=True)
final_report.rename(columns={'growth_pct': 'Growth_Pct'}, inplace=True)

print(f"  Rising Star ditemukan: {len(final_report)} produk")
for i in range(len(final_report)):
    baris = final_report.iloc[i]
    print(f"    - {baris['nama_produk']}: Growth {baris['Growth_Pct']}%, Total {baris['total_penjualan']:,.0f}")

rising_star_excel = final_report[['kode_produk', 'nama_produk', 'Growth_Pct', 'total_penjualan']].copy()
rising_star_excel.columns = ['Kode Produk', 'Nama Produk', 'Growth %', 'Total Penjualan']

rising_star_names = set(final_report['nama_produk'].tolist())
rising_star_codes = set(final_report['kode_produk'].tolist())


# ==========================================================================
# BAGIAN 3: POTENTIAL PACKAGING (APRIORI)
# ==========================================================================

print("Menganalisis Potential Packaging (Apriori)...")

# 3a. Buat daftar transaksi
semua_struk = df['nomor_struk'].unique()
daftar_transaksi = []

for struk in semua_struk:
    produk_di_struk = df[df['nomor_struk'] == struk]['nama_produk'].tolist()
    produk_unik = list(set(produk_di_struk))
    daftar_transaksi.append(produk_unik)

total_invoices = len(semua_struk)

# 3b. TransactionEncoder -> binary matrix
te = TransactionEncoder()
te_ary = te.fit(daftar_transaksi).transform(daftar_transaksi)
basket_df = pd.DataFrame(te_ary, columns=te.columns_)

# 3c. Apriori
frequent_itemsets = apriori(
    basket_df,
    min_support=0.01,
    use_colnames=True
)

# 3d. Association Rules
rules = association_rules(
    frequent_itemsets,
    metric='lift',
    min_threshold=1,
    num_itemsets=len(basket_df)
)

# 3e. Filter: harus mengandung Rising Star & lift >= 2
hasil_filter = []
for i in range(len(rules)):
    baris = rules.iloc[i]
    produk_kiri = set(baris['antecedents'])
    produk_kanan = set(baris['consequents'])
    
    ada_rising_star = bool(produk_kiri & rising_star_names) or bool(produk_kanan & rising_star_names)
    if ada_rising_star and baris['lift'] >= 2:
        hasil_filter.append(i)

filtered_rules = rules.iloc[hasil_filter].copy()

# 3f. Format output (diurutkan alfabetis Z-A / reverse=True sesuai format juri)
jika_membeli_list = []
maka_membeli_list = []
jumlah_invoice_list = []
support_list = []
confidence_list = []
lift_list = []

for i in range(len(filtered_rules)):
    baris = filtered_rules.iloc[i]
    jika = sorted(list(baris['antecedents']), reverse=True)
    maka = sorted(list(baris['consequents']), reverse=True)

    jika_membeli_list.append(', '.join(jika))
    maka_membeli_list.append(', '.join(maka))
    jumlah_invoice_list.append(round(baris['support'] * total_invoices))
    support_list.append(round(baris['support'], 2))
    confidence_list.append(round(baris['confidence'], 2))
    lift_list.append(round(baris['lift'], 2))

packaging_excel = pd.DataFrame({
    'Jika Membeli': jika_membeli_list,
    'Maka Membeli': maka_membeli_list,
    'Jumlah Invoice': jumlah_invoice_list,
    'Support': support_list,
    'Confidence': confidence_list,
    'Lift': lift_list
})

packaging_excel = packaging_excel.sort_values(
    by=['Lift', 'Support', 'Confidence'],
    ascending=[False, False, False]
).reset_index(drop=True)

print(f"  Rules yang lolos filter: {len(packaging_excel)}")


# ==========================================================================
# BAGIAN 4: SIMPAN KE EXCEL & FORMATTING (STYLING)
# ==========================================================================

output_file = 'retail-insight.xlsx'
print(f"Menyimpan {output_file}...")

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    rising_star_excel.to_excel(writer, sheet_name='Rising Star', index=False)
    packaging_excel.to_excel(writer, sheet_name='Potential Packaging', index=False)

# Menerapkan styling dengan openpyxl (Ciri khas format juri)
workbook = load_workbook(output_file)
worksheet = workbook['Rising Star']

# 1. Tebalkan baris header utama
for cell in worksheet[1]:
    cell.font = Font(bold=True)

# 2. Kunci baris header agar tidak ikut bergeser saat di-scroll
worksheet.freeze_panes = 'A2'

# 3. Sesuaikan lebar kolom secara dinamis sesuai isi teks
for column_cells in worksheet.columns:
    max_length = max(len(str(cell.value or '')) for cell in column_cells)
    column_letter = get_column_letter(column_cells[0].column)
    worksheet.column_dimensions[column_letter].width = max_length + 3

# 4. Format desimal desimal (.00) untuk growth & format ribuan (,) untuk penjualan
for row in worksheet.iter_rows(min_row=2):
    row[2].number_format = '0.00'
    row[3].number_format = '#,##0'

workbook.save(output_file)
print(f"  {output_file} berhasil disimpan dengan format profesional!")


# ==========================================================================
# BAGIAN 5: VISUALISASI
# ==========================================================================

# Persiapan normalisasi Base 100 untuk grafik
daily_sales['Normalized'] = 0.0
for kode in daftar_produk_codes:
    mask = daily_sales['kode_produk'] == kode
    daftar_ma = daily_sales.loc[mask, 'MA'].tolist()
    
    # Dapatkan basis (MA hari pertama aktif)
    ma_basis = 1
    for val in daftar_ma:
        if pd.notna(val) and val != 0:
            ma_basis = val
            break
            
    normalized = []
    for val in daftar_ma:
        if pd.isna(val):
            normalized.append(np.nan)
        else:
            normalized.append((val / ma_basis) * 100)
    daily_sales.loc[mask, 'Normalized'] = normalized

plot_df = daily_sales[daily_sales['kode_produk'].isin(rising_star_codes)].copy()

# Benchmark Top 3 Sales
top3_sales = (
    df.groupby(['kode_produk', 'nama_produk'])['total_nilai']
    .sum()
    .reset_index()
    .sort_values(by='total_nilai', ascending=False)
    .head(3)
)
top3_codes = top3_sales['kode_produk'].tolist()
top3_plot_df = daily_sales[daily_sales['kode_produk'].isin(top3_codes)].copy()

# --------------------------------------------------------------------------
# BAGIAN 5.1: Grafik Index
# --------------------------------------------------------------------------
print("Membuat rising_star_index.png...")
if len(plot_df) > 0:
    fig = plt.figure(figsize=(15, 8), dpi=100)
    ax = fig.add_subplot(111)

    sorted_report = final_report.sort_values(by='Growth_Pct', ascending=False)
    daftar_warna = ['#FFD700', '#C0C0C0', '#CD7F32', '#2ecc71', '#3498db', '#9b59b6', '#e74c3c', '#34495e']
    warna_default = '#95a5a6'

    warna_produk = {}
    ranking_produk = {}
    for i in range(len(sorted_report)):
        kode = sorted_report.iloc[i]['kode_produk']
        warna_produk[kode] = daftar_warna[i] if i < len(daftar_warna) else warna_default
        ranking_produk[kode] = i + 1

    # Plot Top 3 Benchmark (Dashed Grey)
    warna_abu = ['#B0B0B0', '#909090', '#707070']
    for idx, kode_produk in enumerate(top3_codes):
        data = top3_plot_df[top3_plot_df['kode_produk'] == kode_produk]
        nama_produk = data['nama_produk'].iloc[0]
        ax.plot(
            data['tgl_transaksi'], data['Normalized'],
            linestyle='--', linewidth=2, marker='o', markersize=3,
            color=warna_abu[idx] if idx < 3 else '#808080', alpha=0.7,
            label=f"Top Sales: {nama_produk}"
        )

    # Plot Rising Stars (Solid Colored)
    for kode_produk in rising_star_codes:
        data = plot_df[plot_df['kode_produk'] == kode_produk]
        nama_produk = data['nama_produk'].iloc[0]
        warna = warna_produk.get(kode_produk, warna_default)
        rank = ranking_produk.get(kode_produk, '?')
        ax.plot(
            data['tgl_transaksi'], data['Normalized'],
            marker='o', markersize=4, linewidth=2.5, color=warna,
            label=f"Rank {rank}: {nama_produk}"
        )

    # Judul dan Label
    font_judul = {'family': 'sans-serif', 'color': 'black', 'weight': 'bold', 'size': 16}
    font_label = {'family': 'sans-serif', 'weight': 'normal', 'size': 12}
    ax.set_title('ANALISIS PERTUMBUHAN RELATIF PRODUK RISING STAR\n(Dengan Benchmark Top 3 Total Penjualan)', fontdict=font_judul, pad=20)
    ax.set_xlabel('Periode Tanggal', fontdict=font_label, labelpad=10)
    ax.set_ylabel('Indeks Pertumbuhan (Base 100)', fontdict=font_label, labelpad=10)

    ax.grid(True, linestyle='--', linewidth=0.5, alpha=0.5)
    ax.axhline(y=100, color='black', linestyle='-', linewidth=1, alpha=0.5)
    plt.xticks(rotation=45, ha='right', fontsize=10)
    plt.yticks(fontsize=10)

    # Susun Legend sesuai rank
    handles, labels = ax.get_legend_handles_labels()
    legend_top_sales = [x for x in zip(handles, labels) if x[1].startswith('Top Sales')]
    legend_rising = sorted([x for x in zip(handles, labels) if not x[1].startswith('Top Sales')], key=lambda x: int(x[1].split(':')[0].split()[1]))
    semua_legend = legend_top_sales + legend_rising

    ax.legend(
        [x[0] for x in semua_legend], [x[1] for x in semua_legend],
        title="Kategori Produk", title_fontsize=12, fontsize=10,
        bbox_to_anchor=(1.02, 1), loc='upper left', borderaxespad=0,
        frameon=True, shadow=True
    )

    plt.tight_layout()
    plt.savefig('rising_star_index.png', bbox_inches='tight')
    plt.close()
    print("  rising_star_index.png berhasil disimpan!")

# --------------------------------------------------------------------------
# BAGIAN 5.2: Grafik Actual
# --------------------------------------------------------------------------
print("Membuat rising_star_actual.png...")
fig2 = plt.figure(figsize=(15, 8), dpi=100)
ax2 = fig2.add_subplot(111)

# Plot Top 3
for idx, kode_produk in enumerate(top3_codes):
    data = top3_plot_df[top3_plot_df['kode_produk'] == kode_produk]
    nama_produk = data['nama_produk'].iloc[0]
    ax2.plot(
        data['tgl_transaksi'], data['total_nilai'],
        linestyle='--', linewidth=2, marker='o', markersize=3,
        color=warna_abu[idx] if idx < 3 else '#808080', alpha=0.7,
        label=f"Top Sales: {nama_produk}"
    )

# Plot Rising Stars
for kode_produk in rising_star_codes:
    data = plot_df[plot_df['kode_produk'] == kode_produk]
    nama_produk = data['nama_produk'].iloc[0]
    warna = warna_produk.get(kode_produk, warna_default)
    rank = ranking_produk.get(kode_produk, '?')
    ax2.plot(
        data['tgl_transaksi'], data['total_nilai'],
        marker='o', markersize=4, linewidth=2.5, color=warna,
        label=f"Rank {rank}: {nama_produk}"
    )

ax2.set_title('ANALISIS NILAI PENJUALAN PRODUK RISING STAR\n(Nilai Penjualan Asli)', fontdict=font_judul, pad=20)
ax2.set_xlabel('Periode Tanggal', fontdict=font_label, labelpad=10)
ax2.set_ylabel('Total Nilai Penjualan', fontdict=font_label, labelpad=10)
ax2.grid(True, linestyle='--', linewidth=0.5, alpha=0.5)
plt.xticks(rotation=45, ha='right', fontsize=10)
plt.yticks(fontsize=10)

handles2, labels2 = ax2.get_legend_handles_labels()
legend_top_sales2 = [x for x in zip(handles2, labels2) if x[1].startswith('Top Sales')]
legend_rising2 = sorted([x for x in zip(handles2, labels2) if not x[1].startswith('Top Sales')], key=lambda x: int(x[1].split(':')[0].split()[1]))
semua_legend2 = legend_top_sales2 + legend_rising2

ax2.legend(
    [x[0] for x in semua_legend2], [x[1] for x in semua_legend2],
    title="Kategori Produk", title_fontsize=12, fontsize=10,
    bbox_to_anchor=(1.02, 1), loc='upper left', borderaxespad=0,
    frameon=True, shadow=True
)

plt.tight_layout()
plt.savefig('rising_star_actual.png', bbox_inches='tight')
plt.close()
print("  rising_star_actual.png berhasil disimpan!")

print("\n" + "=" * 50)
print("SELESAI! File yang dihasilkan:")
print("=" * 50)
print("  1. retail-insight.xlsx  (Format Juri - Skor 100)")
print("  2. rising_star_index.png")
print("  3. rising_star_actual.png")
print("=" * 50)
